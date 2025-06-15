import os
import zipfile
import rarfile
import requests
import tempfile
from PIL import Image
import sys
from pdf2image import convert_from_path
import openpyxl
import subprocess
import xlrd
import base64
from groq import Groq
import shutil
import json

client = Groq(api_key="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")

"""
Organization Type Enums:
1. Item-rate: Traditional contract where payment is made based on measured quantities of work
2. EPC (Engineering, Procurement, Construction): Turnkey contract where contractor handles all aspects
3. HAM (Hybrid Annuity Model): Public-private partnership with 40% government funding
4. BOT (Build-Operate-Transfer): Private entity builds, operates for concession period, then transfers
"""

TENDER_TYPES = ["item-rate", "epc", "ham", "bot"]

def validate_tender_type(tender_type):
    """Validate and normalize tender type"""
    tender_type = tender_type.lower().strip()
    for valid_type in TENDER_TYPES:
        if valid_type in tender_type:
            return valid_type
    return None

def analyze_with_groq(file_path, answers):
    """Analyze document with Groq's Llama model to extract required info"""
    try:
        if file_path.lower().endswith(('.png', '.jpg', '.jpeg')):
            with open(file_path, "rb") as image_file:
                encoded_image = base64.b64encode(image_file.read()).decode('utf-8')
                image_url = f"data:image/png;base64,{encoded_image}"

                completion = client.chat.completions.create(
                    model="meta-llama/llama-4-maverick-17b-128e-instruct",
                    messages=[
                        {
                            "role": "user",
                            "content": [
                                {
                                    "type": "text",
                                    "text": "Extract these exact details from the tender document:\n"
                                            "1. the Length of road (KM) to be worked on - only numbers like '5.2'\n"
                                            "2. Type of Tender - must be one of: item-rate, epc, ham, bot\n"
                                            "3. Road location name - only official name from document\n"
                                            "Return ONLY as valid JSON with these keys: "
                                            "'length_of_road', 'tender_type', 'road_location'"
                                },
                                {
                                    "type": "image_url",
                                    "image_url": {
                                        "url": image_url
                                    }
                                }
                            ]
                        }
                    ],
                    temperature=0.1,  # Low temperature for precise answers
                    max_tokens=200,
                    response_format={"type": "json_object"}
                )

        elif file_path.lower().endswith('.txt'):
            with open(file_path, 'r') as f:
                text_content = f.read()

                completion = client.chat.completions.create(
                    model="meta-llama/llama-4-maverick-17b-128e-instruct",
                    messages=[
                        {
                            "role": "user",
                            "content": f"Extract these exact details from the tender document:\n"
                                      "1. Length of road (KM) to be worked on - only numbers like '5.2'\n"
                                      "2. Type of Tender - must be one of: item-rate, epc, ham, bot\n"
                                      "3. Road location name - only official name from document\n"
                                      "Document content:\n{text_content}\n"
                                      "Return ONLY as valid JSON with these keys: "
                                      "'length_of_road', 'tender_type', 'road_location'"
                        }
                    ],
                    temperature=0.1,
                    max_tokens=200,
                    response_format={"type": "json_object"}
                )

        response = completion.choices[0].message.content
        print(f"Analysis for {file_path}: {response}")

        try:
            data = json.loads(response)
            updated = False


            if "length_of_road" in data and not answers.get("length_of_road"):
                try:
                    if data["length_of_road"] is not None:
                        answers["length_of_road"] = float(data["length_of_road"])
                        updated = True
                except (ValueError, TypeError):
                    pass

            if "tender_type" in data and not answers.get("tender_type"):
                tender_type = validate_tender_type(data["tender_type"])
                if tender_type:
                    answers["tender_type"] = tender_type
                    updated = True

            if "road_location" in data and not answers.get("road_location"):
                answers["road_location"] = data["road_location"].strip()
                updated = True

            if updated:
                print(f"Updated answers: {answers}")

            # Check if we have all answers
            if all(k in answers for k in ["length_of_road", "tender_type", "road_location"]):
                print("\n✅ All answers found!")
                return answers

        except json.JSONDecodeError:
            print("Invalid JSON response")

    except Exception as e:
        print(f"Error analyzing with Groq: {e}")

    return answers

def process_pdf(file_path, answers):
    """Process PDF file - convert first 2 pages to images and analyze"""
    try:
        images = convert_from_path(file_path, first_page=1, last_page=2, poppler_path = "C:\Program Files\poppler-24.08.0\Library\bin")
        for i, image in enumerate(images):
            screenshot_path = f"{os.path.splitext(file_path)[0]}_page_{i+1}.png"
            image.save(screenshot_path, 'PNG')
            if analyze_with_groq(screenshot_path, answers):
                return answers
    except Exception as e:
        print(f"Error processing PDF: {e}")
    return answers

def process_image(file_path, answers):
    """Process image file and analyze"""
    try:
        if analyze_with_groq(file_path, answers):
            return answers
    except Exception as e:
        print(f"Error processing image: {e}")
    return answers

def process_word(file_path, answers):
    """Convert DOC/DOCX to PDF and process"""
    try:
        pdf_path = f"{os.path.splitext(file_path)[0]}.pdf"
        result = subprocess.run(["unoconv", "-f", "pdf", "-o", pdf_path, file_path],
                              capture_output=True, text=True)

        if result.returncode == 0:
            return process_pdf(pdf_path, answers)
    except Exception as e:
        print(f"Error processing Word file: {e}")
    return answers

def process_excel(file_path, answers):
    """Process Excel file and analyze - first 10000 tokens only"""
    try:
        ext = os.path.splitext(file_path)[1].lower()

        if ext == ".xlsx":
            wb = openpyxl.load_workbook(file_path)
            for i, sheetname in enumerate(wb.sheetnames[:2]):  # First 2 sheets only
                sheet = wb[sheetname]
                screenshot_path = f"{os.path.splitext(file_path)[0]}_sheet_{i+1}.txt"

                # Extract first 10000 tokens
                tokens = []
                for row in sheet.iter_rows(values_only=True):
                    row_tokens = ' '.join(map(str, row)).split()
                    tokens.extend(row_tokens)
                    if len(tokens) >= 10000:
                        tokens = tokens[:10000]
                        break

                # Save first 10000 tokens
                with open(screenshot_path, 'w') as f:
                    f.write(' '.join(tokens))

                print(f"Saved first 10000 tokens: {screenshot_path}")
                if analyze_with_groq(screenshot_path, answers):
                    return answers

        elif ext == ".xls":
            wb = xlrd.open_workbook(file_path)
            for i, sheet in enumerate(wb.sheets()[:2]):  # First 2 sheets only
                screenshot_path = f"{os.path.splitext(file_path)[0]}_sheet_{i+1}.txt"

                # Extract first 10000 tokens
                tokens = []
                for row_idx in range(sheet.nrows):
                    row_tokens = ' '.join(map(str, sheet.row_values(row_idx))).split()
                    tokens.extend(row_tokens)
                    if len(tokens) >= 10000:
                        tokens = tokens[:10000]
                        break

                # Save first 10000 tokens
                with open(screenshot_path, 'w') as f:
                    f.write(' '.join(tokens))

                print(f"Saved first 10000 tokens: {screenshot_path}")
                if analyze_with_groq(screenshot_path, answers):
                    return answers

    except Exception as e:
        print(f"Error processing Excel file: {e}")
    return answers

def process_files(root_dir):
    """Process all supported files in directory"""
    answers = {}
    supported_extensions = ('.pdf', '.jpg', '.jpeg', '.doc', '.docx', '.xls', '.xlsx')

    for root, _, files in os.walk(root_dir):
        for file in files:
            file_path = os.path.join(root, file)
            ext = os.path.splitext(file)[1].lower()

            if ext in supported_extensions:
                try:
                    if ext == '.pdf' and process_pdf(file_path, answers):
                        return answers
                    elif ext in ('.jpg', '.jpeg') and process_image(file_path, answers):
                        return answers
                    elif ext in ('.doc', '.docx') and process_word(file_path, answers):
                        return answers
                    elif ext in ('.xls', '.xlsx') and process_excel(file_path, answers):
                        return answers
                except Exception as e:
                    print(f"Error processing {file_path}: {e}")

    return answers
